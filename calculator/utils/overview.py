import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment, Font,  PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def create_overview_sheet(output_file_path):
    workbook = load_workbook(output_file_path)
    
    results_df = pd.read_excel(output_file_path, sheet_name=workbook.sheetnames[0])
    
    if 'Overview' in workbook.sheetnames:
        overview_sheet = workbook['Overview']
    else:
        overview_sheet = workbook.create_sheet(title='Overview')
    
    total_fees_per_country = results_df.groupby('Publication Country')['Total Fees'].sum().reset_index()
    
    year_columns = [col for col in results_df.columns if col.isdigit()]
    total_fees_per_year = results_df[year_columns].sum().reset_index()
    total_fees_per_year.columns = ['Year', 'Maintenance Cost ($)']

    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center')


    for row in total_fees_per_country.itertuples(index=False):
        overview_sheet.append(row)

    overview_sheet.append([])

    overview_sheet.append(['Year', 'Maintenance Cost ($)'])
    for cell in overview_sheet[overview_sheet.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row in total_fees_per_year.itertuples(index=False):
        overview_sheet.append(row)
    
    main_sheet = workbook[workbook.sheetnames[0]]
    for col in range(1, overview_sheet.max_column + 1):
        col_letter = get_column_letter(col)
        overview_sheet.column_dimensions[col_letter].width = 15 

    overview_sheet.row_dimensions[1].height = main_sheet.row_dimensions[1].height
    overview_sheet.row_dimensions[overview_sheet.max_row - len(total_fees_per_year) - 1].height = main_sheet.row_dimensions[1].height

    for col in range(1, overview_sheet.max_column + 1):
        col_letter = get_column_letter(col)
        overview_sheet.column_dimensions[col_letter].width = 15

    workbook.save(output_file_path)
    print(f"Overview sheet added to {output_file_path}")


def format_dates_and_currency(output_file_path):
    workbook = load_workbook(output_file_path)
    main_sheet = workbook[workbook.sheetnames[0]]

    date_style = NamedStyle(name='short_date', number_format='MM/DD/YYYY')
    currency_style = NamedStyle(name='currency', number_format='$#,##0.00')
    alignment_style = Alignment(horizontal='center', vertical='top')
    header_alignment_style = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)

    date_columns = ['C', 'D', 'E', 'F']
    for col in date_columns:
        for cell in main_sheet[col][1:]:
            cell.style = date_style

    for col in range(11, main_sheet.max_column + 1): 
        col_letter = get_column_letter(col)
        for cell in main_sheet[col_letter][1:]: 
            cell.style = currency_style

    for col in range(1, main_sheet.max_column + 1):
        col_letter = get_column_letter(col)
        if col_letter in date_columns:
            main_sheet.column_dimensions[col_letter].width = 15
        else:
            main_sheet.column_dimensions[col_letter].width = 15

    main_sheet.row_dimensions[1].height = 30

    main_sheet.sheet_view.zoomScale = 80

    for row in main_sheet.iter_rows():
        for cell in row:
            if cell.row == 1:
                cell.alignment = header_alignment_style
                cell.font = bold_font
            else:
                cell.alignment = alignment_style

    main_sheet.title = 'Fees per Year'

    if 'Overview' in workbook.sheetnames:
        overview_sheet = workbook['Overview']
        for cell in overview_sheet['B'][1:]:
            cell.style = currency_style
            
        for col in range(1, overview_sheet.max_column + 1):
            col_letter = get_column_letter(col)
            overview_sheet.column_dimensions[col_letter].width = 20

        overview_sheet.row_dimensions[1].height = 30

        for row in overview_sheet.iter_rows():
            for cell in row:
                if cell.row == 1:
                    cell.alignment = header_alignment_style
                    cell.font = bold_font
                else:
                    cell.alignment = alignment_style

    workbook.save(output_file_path)
